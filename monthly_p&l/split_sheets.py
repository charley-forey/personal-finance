"""Split each tab from Personal Monthly P&L.xlsx into its own workbook."""

from pathlib import Path

from openpyxl import load_workbook

FOLDER = Path(__file__).parent
SOURCE = FOLDER / "Personal Monthly P&L.xlsx"


def safe_filename(name: str) -> str:
    for char in '<>:"/\\|?*':
        name = name.replace(char, "-")
    return name.strip()


def main() -> None:
    wb = load_workbook(SOURCE, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    print(f"Splitting {len(sheet_names)} sheets from {SOURCE.name}:\n")

    for sheet_name in sheet_names:
        temp_wb = load_workbook(SOURCE)
        for other in list(temp_wb.sheetnames):
            if other != sheet_name:
                del temp_wb[other]

        out_path = FOLDER / f"{safe_filename(sheet_name)}.xlsx"
        temp_wb.save(out_path)
        temp_wb.close()
        print(f"  -> {out_path.name}")

    print("\nDone.")


if __name__ == "__main__":
    main()
